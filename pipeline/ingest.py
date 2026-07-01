"""
Ingestion — load the OG .xlsx (or a CSV) into normalized Record objects.

Handles the confidential banner (real header at row 6), numeric scales, and free-text
severity words. Missing values stay ABSENT (not zero) so completeness is meaningful.
"""

from __future__ import annotations
import os
from typing import List, Optional, Dict, Any

from .schema import (META_COLS, QUAD_COLUMNS, TARGET_COLS, NUMERIC_SCALES,
                     IMAGE_COL, HEADER_ROW, all_feature_columns)
from .core import Record

# free-text severity lexicon -> 0..1
SEVERITY_WORDS = {
    "none": 0.0, "nil": 0.0, "no": 0.0, "absent": 0.0, "clear": 0.0,
    "minimal": 0.15, "trace": 0.15, "very mild": 0.2,
    "mild": 0.33, "low": 0.33, "slight": 0.3,
    "moderate": 0.6, "medium": 0.6,
    "high": 0.85, "marked": 0.8, "severe": 1.0, "extreme": 1.0, "very high": 0.95,
    "yes": 1.0, "present": 0.7,
}


def _normalize(col: str, value: Any) -> Optional[float]:
    """Coerce a cell to a 0..1 float, or None if truly missing/unusable."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.lower() in ("nan", "n/a", "na", "none listed"):
        return None
    # explicit numeric scale
    if col in NUMERIC_SCALES:
        try:
            return max(0.0, min(1.0, float(s) / NUMERIC_SCALES[col]))
        except ValueError:
            pass
    # plain number -> assume already 0..1-ish, else squash counts with log
    try:
        v = float(s)
        if 0.0 <= v <= 1.0:
            return v
        if v <= 10:
            return min(1.0, v / 10.0)
        import math
        return min(1.0, math.log1p(v) / math.log1p(100))
    except ValueError:
        pass
    # free-text severity
    key = s.lower()
    if key in SEVERITY_WORDS:
        return SEVERITY_WORDS[key]
    for word, val in SEVERITY_WORDS.items():
        if word in key:
            return val
    return 0.5  # unknown non-empty categorical -> mid signal (present but unscored)


def _rows_from_xlsx(path: str):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[HEADER_ROW - 1]]
    for r in rows[HEADER_ROW:]:
        if r is None or all(c in (None, "") for c in r):
            continue
        yield dict(zip(header, r))


def _rows_from_csv(path: str):
    import csv
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            yield row


def load_records(path: str) -> List[Record]:
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    reader = _rows_from_xlsx if path.lower().endswith((".xlsx", ".xlsm")) else _rows_from_csv
    feature_cols = all_feature_columns()
    records: List[Record] = []
    for row in reader(path):
        img = row.get(IMAGE_COL)
        img = str(img).strip() if img not in (None, "") else None
        feats: Dict[str, float] = {}
        for c in feature_cols:
            v = _normalize(c, row.get(c))
            if v is not None:
                feats[c] = v
        rec = Record(
            pic_no=row.get("PicNo"),
            image_file=img,
            meta={k: row.get(k) for k in META_COLS if k in row},
            features=feats,
            raw={k: row.get(k) for k in feature_cols if k in row},
            targets={k: row.get(k) for k in TARGET_COLS if k in row},
        )
        records.append(rec)
    return records
